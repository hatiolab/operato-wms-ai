#!/usr/bin/env python3
"""
OMS 주문 임포트용 샘플 엑셀 생성 스크립트

ImportShipmentOrder 엔티티 기준으로 B2C/B2B 샘플 데이터가 포함된 엑셀 파일을 생성한다.
Excel 컬럼 헤더는 snake_case (Jackson SNAKE_CASE 설정에 맞춤)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 공통 스타일 ───
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# ─── 컬럼 정의 (header, width, required) ───
COLUMNS = [
    ("ref_order_no",      20, True),
    ("order_date",        14, False),
    ("ship_by_date",      14, False),
    ("com_cd",            12, False),
    ("cust_cd",           12, False),
    ("cust_nm",           16, False),
    ("wh_cd",             10, False),
    ("biz_type",          12, False),
    ("ship_type",         14, False),
    ("dlv_type",          14, False),
    ("priority_cd",       12, False),
    ("line_no",            8, False),
    ("sku_cd",            14, True),
    ("sku_nm",            20, False),
    ("order_qty",         10, True),
    ("unit_price",        10, False),
    ("barcode",           16, False),
    ("lot_no",            14, False),
    ("expired_date",      14, False),
    ("sender_nm",         14, False),
    ("sender_phone",      16, False),
    ("sender_zip_cd",     12, False),
    ("sender_addr",       30, False),
    ("orderer_nm",        14, False),
    ("receiver_nm",       14, False),
    ("receiver_phone",    16, False),
    ("receiver_zip_cd",   12, False),
    ("receiver_addr",     30, False),
    ("receiver_addr2",    20, False),
    ("delivery_memo",     24, False),
    ("remarks",           20, False),
]

# ─── B2C 샘플 데이터 (5주문, 10행) ───
B2C_DATA = [
    # 주문1: 단품 (1SKU)
    {
        "ref_order_no": "ORD-20260329-001", "order_date": "2026-03-29", "ship_by_date": "2026-03-31",
        "com_cd": "ACME", "cust_cd": "CH-NAVER", "cust_nm": "네이버스토어",
        "wh_cd": "WH-01", "biz_type": "B2C_OUT", "ship_type": "PARCEL", "dlv_type": "STANDARD",
        "priority_cd": "NORMAL", "line_no": "1",
        "sku_cd": "SKU-A001", "sku_nm": "무선 블루투스 이어폰", "order_qty": 1, "unit_price": 35000,
        "barcode": "8801234567890", "lot_no": "", "expired_date": "",
        "sender_nm": "ACME물류", "sender_phone": "02-1234-5678", "sender_zip_cd": "06234",
        "sender_addr": "서울시 강남구 테헤란로 123",
        "orderer_nm": "김민수",
        "receiver_nm": "김민수", "receiver_phone": "010-1234-5678", "receiver_zip_cd": "13529",
        "receiver_addr": "경기도 성남시 분당구 판교로 456", "receiver_addr2": "판교테크노밸리 A동 301호",
        "delivery_memo": "부재시 경비실에 맡겨주세요", "remarks": "",
    },
    # 주문2: 다품 (3SKU, 동일 ref_order_no)
    {
        "ref_order_no": "ORD-20260329-002", "order_date": "2026-03-29", "ship_by_date": "2026-03-30",
        "com_cd": "ACME", "cust_cd": "CH-COUPANG", "cust_nm": "쿠팡",
        "wh_cd": "WH-01", "biz_type": "B2C_OUT", "ship_type": "PARCEL", "dlv_type": "EXPRESS",
        "priority_cd": "HIGH", "line_no": "1",
        "sku_cd": "SKU-B002", "sku_nm": "프리미엄 텀블러 500ml", "order_qty": 2, "unit_price": 28000,
        "barcode": "8801234567891", "lot_no": "", "expired_date": "",
        "sender_nm": "ACME물류", "sender_phone": "02-1234-5678", "sender_zip_cd": "06234",
        "sender_addr": "서울시 강남구 테헤란로 123",
        "orderer_nm": "이지은",
        "receiver_nm": "이지은", "receiver_phone": "010-2345-6789", "receiver_zip_cd": "04524",
        "receiver_addr": "서울시 중구 명동길 12", "receiver_addr2": "3층",
        "delivery_memo": "로켓배송", "remarks": "",
    },
    {
        "ref_order_no": "ORD-20260329-002", "order_date": "2026-03-29", "ship_by_date": "2026-03-30",
        "com_cd": "ACME", "cust_cd": "CH-COUPANG", "cust_nm": "쿠팡",
        "wh_cd": "WH-01", "biz_type": "B2C_OUT", "ship_type": "PARCEL", "dlv_type": "EXPRESS",
        "priority_cd": "HIGH", "line_no": "2",
        "sku_cd": "SKU-C003", "sku_nm": "USB-C 충전 케이블 1m", "order_qty": 3, "unit_price": 8900,
        "barcode": "8801234567892", "lot_no": "", "expired_date": "",
        "sender_nm": "", "sender_phone": "", "sender_zip_cd": "",
        "sender_addr": "",
        "orderer_nm": "",
        "receiver_nm": "", "receiver_phone": "", "receiver_zip_cd": "",
        "receiver_addr": "", "receiver_addr2": "",
        "delivery_memo": "", "remarks": "",
    },
    {
        "ref_order_no": "ORD-20260329-002", "order_date": "2026-03-29", "ship_by_date": "2026-03-30",
        "com_cd": "ACME", "cust_cd": "CH-COUPANG", "cust_nm": "쿠팡",
        "wh_cd": "WH-01", "biz_type": "B2C_OUT", "ship_type": "PARCEL", "dlv_type": "EXPRESS",
        "priority_cd": "HIGH", "line_no": "3",
        "sku_cd": "SKU-D004", "sku_nm": "노트북 파우치 15인치", "order_qty": 1, "unit_price": 22000,
        "barcode": "8801234567893", "lot_no": "", "expired_date": "",
        "sender_nm": "", "sender_phone": "", "sender_zip_cd": "",
        "sender_addr": "",
        "orderer_nm": "",
        "receiver_nm": "", "receiver_phone": "", "receiver_zip_cd": "",
        "receiver_addr": "", "receiver_addr2": "",
        "delivery_memo": "", "remarks": "",
    },
    # 주문3: 새벽배송 주문
    {
        "ref_order_no": "ORD-20260329-003", "order_date": "2026-03-29", "ship_by_date": "2026-03-29",
        "com_cd": "ACME", "cust_cd": "CH-SSG", "cust_nm": "SSG닷컴",
        "wh_cd": "WH-01", "biz_type": "B2C_OUT", "ship_type": "PARCEL", "dlv_type": "DAWN",
        "priority_cd": "URGENT", "line_no": "1",
        "sku_cd": "SKU-E005", "sku_nm": "유기농 그래놀라 500g", "order_qty": 2, "unit_price": 15000,
        "barcode": "8801234567894", "lot_no": "LOT-2026-001", "expired_date": "2026-09-30",
        "sender_nm": "ACME물류", "sender_phone": "02-1234-5678", "sender_zip_cd": "06234",
        "sender_addr": "서울시 강남구 테헤란로 123",
        "orderer_nm": "박서연",
        "receiver_nm": "박서연", "receiver_phone": "010-3456-7890", "receiver_zip_cd": "06035",
        "receiver_addr": "서울시 강남구 역삼동 789-12", "receiver_addr2": "힐스테이트 1402호",
        "delivery_memo": "새벽배송 - 문 앞에 놓아주세요", "remarks": "신선식품 주의",
    },
    # 주문4: 당일배송 다품
    {
        "ref_order_no": "ORD-20260329-004", "order_date": "2026-03-29", "ship_by_date": "2026-03-29",
        "com_cd": "ACME", "cust_cd": "CH-NAVER", "cust_nm": "네이버스토어",
        "wh_cd": "WH-01", "biz_type": "B2C_OUT", "ship_type": "PARCEL", "dlv_type": "SAME_DAY",
        "priority_cd": "HIGH", "line_no": "1",
        "sku_cd": "SKU-F006", "sku_nm": "스마트워치 밴드 (블랙)", "order_qty": 1, "unit_price": 19900,
        "barcode": "8801234567895", "lot_no": "", "expired_date": "",
        "sender_nm": "ACME물류", "sender_phone": "02-1234-5678", "sender_zip_cd": "06234",
        "sender_addr": "서울시 강남구 테헤란로 123",
        "orderer_nm": "최준혁",
        "receiver_nm": "최준혁", "receiver_phone": "010-4567-8901", "receiver_zip_cd": "05510",
        "receiver_addr": "서울시 송파구 올림픽로 300", "receiver_addr2": "잠실엘스 2023호",
        "delivery_memo": "", "remarks": "",
    },
    {
        "ref_order_no": "ORD-20260329-004", "order_date": "2026-03-29", "ship_by_date": "2026-03-29",
        "com_cd": "ACME", "cust_cd": "CH-NAVER", "cust_nm": "네이버스토어",
        "wh_cd": "WH-01", "biz_type": "B2C_OUT", "ship_type": "PARCEL", "dlv_type": "SAME_DAY",
        "priority_cd": "HIGH", "line_no": "2",
        "sku_cd": "SKU-A001", "sku_nm": "무선 블루투스 이어폰", "order_qty": 1, "unit_price": 35000,
        "barcode": "8801234567890", "lot_no": "", "expired_date": "",
        "sender_nm": "", "sender_phone": "", "sender_zip_cd": "",
        "sender_addr": "",
        "orderer_nm": "",
        "receiver_nm": "", "receiver_phone": "", "receiver_zip_cd": "",
        "receiver_addr": "", "receiver_addr2": "",
        "delivery_memo": "", "remarks": "",
    },
    # 주문5: 일반배송 단품
    {
        "ref_order_no": "ORD-20260329-005", "order_date": "2026-03-29", "ship_by_date": "2026-04-01",
        "com_cd": "ACME", "cust_cd": "CH-11ST", "cust_nm": "11번가",
        "wh_cd": "WH-01", "biz_type": "B2C_OUT", "ship_type": "PARCEL", "dlv_type": "STANDARD",
        "priority_cd": "LOW", "line_no": "1",
        "sku_cd": "SKU-G007", "sku_nm": "데스크 LED 조명", "order_qty": 1, "unit_price": 45000,
        "barcode": "8801234567896", "lot_no": "", "expired_date": "",
        "sender_nm": "ACME물류", "sender_phone": "02-1234-5678", "sender_zip_cd": "06234",
        "sender_addr": "서울시 강남구 테헤란로 123",
        "orderer_nm": "정하늘",
        "receiver_nm": "정하늘", "receiver_phone": "010-5678-9012", "receiver_zip_cd": "16508",
        "receiver_addr": "경기도 수원시 영통구 광교로 100", "receiver_addr2": "광교마을 5단지 707호",
        "delivery_memo": "택배함에 넣어주세요", "remarks": "",
    },
]

# ─── B2B 샘플 데이터 (3주문, 8행) ───
B2B_DATA = [
    # 주문1: 거래처 납품 (3SKU)
    {
        "ref_order_no": "PO-20260329-001", "order_date": "2026-03-29", "ship_by_date": "2026-04-02",
        "com_cd": "ACME", "cust_cd": "CUST-SAMSUNG", "cust_nm": "삼성전자",
        "wh_cd": "WH-01", "biz_type": "B2B_OUT", "ship_type": "FREIGHT", "dlv_type": "STANDARD",
        "priority_cd": "NORMAL", "line_no": "1",
        "sku_cd": "SKU-H010", "sku_nm": "산업용 커넥터 A타입", "order_qty": 500, "unit_price": 1200,
        "barcode": "8801234568001", "lot_no": "LOT-2026-100", "expired_date": "",
        "sender_nm": "ACME물류센터", "sender_phone": "02-1234-5678", "sender_zip_cd": "06234",
        "sender_addr": "서울시 강남구 테헤란로 123",
        "orderer_nm": "",
        "receiver_nm": "삼성전자 수원사업장", "receiver_phone": "031-200-1234", "receiver_zip_cd": "16677",
        "receiver_addr": "경기도 수원시 영통구 삼성로 129", "receiver_addr2": "자재입고동 B1",
        "delivery_memo": "입고 시간: 09:00~17:00", "remarks": "PO-2026-SM-001",
    },
    {
        "ref_order_no": "PO-20260329-001", "order_date": "2026-03-29", "ship_by_date": "2026-04-02",
        "com_cd": "ACME", "cust_cd": "CUST-SAMSUNG", "cust_nm": "삼성전자",
        "wh_cd": "WH-01", "biz_type": "B2B_OUT", "ship_type": "FREIGHT", "dlv_type": "STANDARD",
        "priority_cd": "NORMAL", "line_no": "2",
        "sku_cd": "SKU-I011", "sku_nm": "산업용 커넥터 B타입", "order_qty": 300, "unit_price": 1500,
        "barcode": "8801234568002", "lot_no": "LOT-2026-101", "expired_date": "",
        "sender_nm": "", "sender_phone": "", "sender_zip_cd": "",
        "sender_addr": "",
        "orderer_nm": "",
        "receiver_nm": "", "receiver_phone": "", "receiver_zip_cd": "",
        "receiver_addr": "", "receiver_addr2": "",
        "delivery_memo": "", "remarks": "",
    },
    {
        "ref_order_no": "PO-20260329-001", "order_date": "2026-03-29", "ship_by_date": "2026-04-02",
        "com_cd": "ACME", "cust_cd": "CUST-SAMSUNG", "cust_nm": "삼성전자",
        "wh_cd": "WH-01", "biz_type": "B2B_OUT", "ship_type": "FREIGHT", "dlv_type": "STANDARD",
        "priority_cd": "NORMAL", "line_no": "3",
        "sku_cd": "SKU-J012", "sku_nm": "케이블 하네스 세트", "order_qty": 200, "unit_price": 3500,
        "barcode": "8801234568003", "lot_no": "LOT-2026-102", "expired_date": "",
        "sender_nm": "", "sender_phone": "", "sender_zip_cd": "",
        "sender_addr": "",
        "orderer_nm": "",
        "receiver_nm": "", "receiver_phone": "", "receiver_zip_cd": "",
        "receiver_addr": "", "receiver_addr2": "",
        "delivery_memo": "", "remarks": "",
    },
    # 주문2: 매장 납품
    {
        "ref_order_no": "PO-20260329-002", "order_date": "2026-03-29", "ship_by_date": "2026-04-01",
        "com_cd": "ACME", "cust_cd": "CUST-LOTTE", "cust_nm": "롯데마트",
        "wh_cd": "WH-01", "biz_type": "B2B_OUT", "ship_type": "DIRECT", "dlv_type": "STANDARD",
        "priority_cd": "NORMAL", "line_no": "1",
        "sku_cd": "SKU-K013", "sku_nm": "프리미엄 텀블러 세트(3P)", "order_qty": 100, "unit_price": 75000,
        "barcode": "8801234568004", "lot_no": "", "expired_date": "",
        "sender_nm": "ACME물류센터", "sender_phone": "02-1234-5678", "sender_zip_cd": "06234",
        "sender_addr": "서울시 강남구 테헤란로 123",
        "orderer_nm": "",
        "receiver_nm": "롯데마트 잠실점", "receiver_phone": "02-411-7777", "receiver_zip_cd": "05551",
        "receiver_addr": "서울시 송파구 올림픽로 240", "receiver_addr2": "잠실점 물류센터",
        "delivery_memo": "매장 입고 시간 06:00~08:00", "remarks": "",
    },
    {
        "ref_order_no": "PO-20260329-002", "order_date": "2026-03-29", "ship_by_date": "2026-04-01",
        "com_cd": "ACME", "cust_cd": "CUST-LOTTE", "cust_nm": "롯데마트",
        "wh_cd": "WH-01", "biz_type": "B2B_OUT", "ship_type": "DIRECT", "dlv_type": "STANDARD",
        "priority_cd": "NORMAL", "line_no": "2",
        "sku_cd": "SKU-L014", "sku_nm": "스테인리스 보온병 750ml", "order_qty": 150, "unit_price": 32000,
        "barcode": "8801234568005", "lot_no": "", "expired_date": "",
        "sender_nm": "", "sender_phone": "", "sender_zip_cd": "",
        "sender_addr": "",
        "orderer_nm": "",
        "receiver_nm": "", "receiver_phone": "", "receiver_zip_cd": "",
        "receiver_addr": "", "receiver_addr2": "",
        "delivery_memo": "", "remarks": "",
    },
    # 주문3: 창고간 이동
    {
        "ref_order_no": "PO-20260329-003", "order_date": "2026-03-29", "ship_by_date": "2026-04-03",
        "com_cd": "ACME", "cust_cd": "CUST-INTERNAL", "cust_nm": "ACME 부산센터",
        "wh_cd": "WH-01", "biz_type": "B2B_OUT", "ship_type": "FREIGHT", "dlv_type": "STANDARD",
        "priority_cd": "LOW", "line_no": "1",
        "sku_cd": "SKU-A001", "sku_nm": "무선 블루투스 이어폰", "order_qty": 1000, "unit_price": 25000,
        "barcode": "8801234567890", "lot_no": "LOT-2026-050", "expired_date": "",
        "sender_nm": "ACME서울센터", "sender_phone": "02-1234-5678", "sender_zip_cd": "06234",
        "sender_addr": "서울시 강남구 테헤란로 123",
        "orderer_nm": "",
        "receiver_nm": "ACME부산센터", "receiver_phone": "051-200-5678", "receiver_zip_cd": "46726",
        "receiver_addr": "부산시 강서구 녹산산업중로 333", "receiver_addr2": "ACME물류센터",
        "delivery_memo": "화물차 입고 - 도크3", "remarks": "창고간 재고 이동",
    },
    {
        "ref_order_no": "PO-20260329-003", "order_date": "2026-03-29", "ship_by_date": "2026-04-03",
        "com_cd": "ACME", "cust_cd": "CUST-INTERNAL", "cust_nm": "ACME 부산센터",
        "wh_cd": "WH-01", "biz_type": "B2B_OUT", "ship_type": "FREIGHT", "dlv_type": "STANDARD",
        "priority_cd": "LOW", "line_no": "2",
        "sku_cd": "SKU-B002", "sku_nm": "프리미엄 텀블러 500ml", "order_qty": 500, "unit_price": 20000,
        "barcode": "8801234567891", "lot_no": "LOT-2026-051", "expired_date": "",
        "sender_nm": "", "sender_phone": "", "sender_zip_cd": "",
        "sender_addr": "",
        "orderer_nm": "",
        "receiver_nm": "", "receiver_phone": "", "receiver_zip_cd": "",
        "receiver_addr": "", "receiver_addr2": "",
        "delivery_memo": "", "remarks": "",
    },
    {
        "ref_order_no": "PO-20260329-003", "order_date": "2026-03-29", "ship_by_date": "2026-04-03",
        "com_cd": "ACME", "cust_cd": "CUST-INTERNAL", "cust_nm": "ACME 부산센터",
        "wh_cd": "WH-01", "biz_type": "B2B_OUT", "ship_type": "FREIGHT", "dlv_type": "STANDARD",
        "priority_cd": "LOW", "line_no": "3",
        "sku_cd": "SKU-G007", "sku_nm": "데스크 LED 조명", "order_qty": 300, "unit_price": 32000,
        "barcode": "8801234567896", "lot_no": "", "expired_date": "",
        "sender_nm": "", "sender_phone": "", "sender_zip_cd": "",
        "sender_addr": "",
        "orderer_nm": "",
        "receiver_nm": "", "receiver_phone": "", "receiver_zip_cd": "",
        "receiver_addr": "", "receiver_addr2": "",
        "delivery_memo": "", "remarks": "",
    },
]


def create_sheet(wb, sheet_name, data):
    """시트 생성 및 데이터 입력"""
    ws = wb.active if sheet_name == wb.sheetnames[0] else wb.create_sheet(title=sheet_name)
    if sheet_name != "B2C 주문":
        ws.title = sheet_name

    # 헤더 행
    for col_idx, (header, width, required) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = REQUIRED_FILL if required else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터 행
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, (header, _, _) in enumerate(COLUMNS, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if header == "order_qty":
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif header == "unit_price":
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(horizontal="left")

    # 필터 설정
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(data) + 1}"
    # 행 높이
    ws.row_dimensions[1].height = 24

    return ws


def create_guide_sheet(wb):
    """가이드 시트 생성"""
    ws = wb.create_sheet(title="입력 가이드")

    guide_header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    guide_header_fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
    guide_font = Font(name="맑은 고딕", size=10)

    headers = ["컬럼명", "필수", "설명", "입력 예시", "비고"]
    widths = [18, 6, 35, 25, 30]
    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = guide_header_font
        cell.fill = guide_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    guide_data = [
        ("ref_order_no",    "Y", "참조 주문번호 (채널 주문번호)", "ORD-20260329-001", "동일 ref_order_no의 행은 하나의 주문으로 묶임"),
        ("order_date",      "",  "주문일자 (YYYY-MM-DD)", "2026-03-29", "미입력 시 당일 날짜 자동 설정"),
        ("ship_by_date",    "",  "출하 기한일 (YYYY-MM-DD)", "2026-03-31", ""),
        ("com_cd",          "",  "회사 코드", "ACME", "미입력 시 기본 회사 사용"),
        ("cust_cd",         "",  "고객/채널 코드", "CH-NAVER", "B2C: 판매채널, B2B: 거래처 코드"),
        ("cust_nm",         "",  "고객/채널 명", "네이버스토어", ""),
        ("wh_cd",           "",  "창고 코드", "WH-01", "미입력 시 DEFAULT"),
        ("biz_type",        "",  "업무 유형", "B2C_OUT", "B2C_OUT / B2B_OUT / B2C_RTN / B2B_RTN"),
        ("ship_type",       "",  "출하 유형", "PARCEL", "PARCEL / FREIGHT / DIRECT / STORE_PICKUP"),
        ("dlv_type",        "",  "배송 유형", "STANDARD", "STANDARD / EXPRESS / SAME_DAY / DAWN"),
        ("priority_cd",     "",  "우선순위", "NORMAL", "URGENT / HIGH / NORMAL / LOW"),
        ("line_no",         "",  "라인 번호", "1", "미입력 시 자동 순번 부여"),
        ("sku_cd",          "Y", "상품 코드 (SKU)", "SKU-A001", "사전 등록된 SKU 코드"),
        ("sku_nm",          "",  "상품명", "무선 블루투스 이어폰", ""),
        ("order_qty",       "Y", "주문 수량", "1", "양수만 허용"),
        ("unit_price",      "",  "단가", "35000", ""),
        ("barcode",         "",  "바코드", "8801234567890", ""),
        ("lot_no",          "",  "로트 번호", "LOT-2026-001", ""),
        ("expired_date",    "",  "유통기한 (YYYY-MM-DD)", "2026-09-30", ""),
        ("sender_nm",       "",  "발송인 이름", "ACME물류", "첫 행에만 입력 (동일 주문)"),
        ("sender_phone",    "",  "발송인 전화번호", "02-1234-5678", ""),
        ("sender_zip_cd",   "",  "발송인 우편번호", "06234", ""),
        ("sender_addr",     "",  "발송인 주소", "서울시 강남구 테헤란로 123", ""),
        ("orderer_nm",      "",  "주문자명", "김민수", ""),
        ("receiver_nm",     "",  "수취인 이름", "김민수", "첫 행에만 입력 (동일 주문)"),
        ("receiver_phone",  "",  "수취인 전화번호", "010-1234-5678", ""),
        ("receiver_zip_cd", "",  "수취인 우편번호", "13529", ""),
        ("receiver_addr",   "",  "수취인 주소", "경기도 성남시 분당구 판교로 456", ""),
        ("receiver_addr2",  "",  "수취인 상세주소", "판교테크노밸리 A동 301호", ""),
        ("delivery_memo",   "",  "배송 메모", "부재시 경비실에 맡겨주세요", ""),
        ("remarks",         "",  "비고", "", ""),
    ]

    req_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    for row_idx, (col_name, req, desc, example, note) in enumerate(guide_data, 2):
        values = [col_name, req, desc, example, note]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = guide_font
            cell.border = THIN_BORDER
            if req == "Y":
                cell.fill = req_fill

    ws.row_dimensions[1].height = 24
    return ws


def main():
    wb = Workbook()

    # B2C 시트
    create_sheet(wb, "B2C 주문", B2C_DATA)
    wb.active.title = "B2C 주문"

    # B2B 시트
    create_sheet(wb, "B2B 주문", B2B_DATA)

    # 가이드 시트
    create_guide_sheet(wb)

    output_path = "/Users/shortstop/Git/operato-wms-ai/templates/shipment-order-import-sample.xlsx"
    wb.save(output_path)
    print(f"생성 완료: {output_path}")
    print(f"  - B2C 주문: {len(B2C_DATA)}행 (5주문)")
    print(f"  - B2B 주문: {len(B2B_DATA)}행 (3주문)")
    print(f"  - 입력 가이드 시트 포함")


if __name__ == "__main__":
    main()
